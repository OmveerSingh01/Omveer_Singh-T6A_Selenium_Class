import pytest
import openpyxl
from openpyxl.workbook import Workbook

wb = openpyxl.Workbook()
sheetName = 'sheet1'

if sheetName in wb.sheetnames:
    ws = wb[sheetName]

else:
    ws = wb.create_sheet(sheetName)

ws['A1'] = 'user'
ws['B1'] = 'password'
# wb.save('sample.xlsx')    # It will save in the local directory

ws.append(['user1','123'])
ws.append(['user2','125'])
ws.append(['user3','128'])
ws.append(['standard_user','secret_sauce'])
ws.append(['user5','132'])

ws.delete_rows(2)

wb.save('sample.xlsx')


for row in ws.iter_rows(values_only=True):
    print(row)
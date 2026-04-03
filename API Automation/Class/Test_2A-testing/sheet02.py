import openpyxl
from click.formatting import iter_rows


def get_test_data():
    wb = openpyxl.load_workbook(r"D:\Selenium_Training_py\Selenium_capgemini\API Automation\Class\sample.xlsx")
    ws = wb['sheet1']


    data =[]

    for i in ws.iter_rows(min_row = 2,values_only=True):
        data.append(i)
    return data


